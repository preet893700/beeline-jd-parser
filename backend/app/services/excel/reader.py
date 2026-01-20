# app/services/excel/reader.py
"""
Excel File Reader
Reads and parses Excel files with multiple sheet support
"""

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import logging
from typing import List, Dict, Any
from io import BytesIO

from app.core.exceptions import ExcelReadError

logger = logging.getLogger(__name__)


class ExcelReader:
    """Excel file reader with robust error handling"""
    
    @staticmethod
    async def read_file(file_content: bytes) -> Dict[str, Any]:
        """
        Read Excel file and return structured data
        Returns: {
            "sheets": [
                {
                    "id": str,  # CRITICAL: Unique sheet ID
                    "name": str,
                    "headers": List[str],
                    "rows": List[List[str]],
                    "total_rows": int
                }
            ]
        }
        """
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(
                BytesIO(file_content),
                read_only=True,
                data_only=True
            )
            
            sheets_data = []
            
            for idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                sheet_data = ExcelReader._parse_sheet(sheet)
                sheets_data.append({
                    "id": f"sheet_{idx}_{sheet_name}",  # CRITICAL: Stable unique ID
                    "name": sheet_name,
                    **sheet_data
                })
            
            workbook.close()
            
            return {"sheets": sheets_data}
            
        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            raise ExcelReadError(f"Cannot read Excel file: {str(e)}")
    
    @staticmethod
    def _parse_sheet(sheet: Worksheet) -> Dict[str, Any]:
        """Parse a single worksheet"""
        rows_data = []
        headers = []
        
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            # Convert row values to strings, handle None
            row_values = [
                str(cell) if cell is not None else ""
                for cell in row
            ]
            
            if idx == 0:
                # First row is headers
                headers = row_values
            else:
                rows_data.append(row_values)
        
        return {
            "headers": headers,
            "rows": rows_data,
            "total_rows": len(rows_data)
        }
    
    @staticmethod
    async def extract_column_data(
        file_content: bytes,
        sheet_name: str,
        column_index: int
    ) -> List[str]:
        """Extract all values from a specific column"""
        try:
            workbook = openpyxl.load_workbook(
                BytesIO(file_content),
                read_only=True,
                data_only=True
            )
            
            if sheet_name not in workbook.sheetnames:
                raise ExcelReadError(f"Sheet '{sheet_name}' not found")
            
            sheet = workbook[sheet_name]
            column_data = []
            
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                # Skip header row
                if idx == 0:
                    continue
                
                # Extract column value
                if column_index < len(row):
                    value = row[column_index]
                    column_data.append(str(value) if value else "")
                else:
                    column_data.append("")
            
            workbook.close()
            
            return column_data
            
        except Exception as e:
            logger.error(f"Failed to extract column data: {e}")
            raise ExcelReadError(f"Cannot extract column: {str(e)}")