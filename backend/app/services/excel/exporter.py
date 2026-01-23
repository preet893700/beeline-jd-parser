# app/services/excel/exporter.py
"""
Excel File Exporter
Creates Excel files with extraction results

UPDATED: Added Min Bill Rate and Max Bill Rate columns
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
import logging
from typing import List
from io import BytesIO

from app.models.jd_result import ExcelJDRow

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export extraction results to Excel"""
    
    @staticmethod
    async def create_result_workbook(
        original_file: bytes,
        sheet_name: str,
        jd_column_index: int,
        results: List[ExcelJDRow]
    ) -> BytesIO:
        """
        Create Excel workbook with extraction results
        Inserts extracted columns after JD column
        """
        try:
            # Load original workbook
            workbook = openpyxl.load_workbook(BytesIO(original_file))
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found")
            
            sheet = workbook[sheet_name]
            
            # Define new column headers (UPDATED: Added Min/Max Bill Rate)
            new_headers = [
                "Bill Rate",
                "Min Bill Rate",  # NEW
                "Max Bill Rate",  # NEW
                "Duration",
                "Experience",
                "GBAMS/RGS ID",
                "Location",
                "Skills",
                "Role Description",
                "MSP Owner"
            ]
            
            # Insert columns after JD column
            insert_at = jd_column_index + 2  # +2 because Excel is 1-indexed
            
            for i in range(len(new_headers)):
                sheet.insert_cols(insert_at + i)
            
            # Add headers
            header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            header_font = Font(bold=True)
            
            for i, header in enumerate(new_headers):
                cell = sheet.cell(row=1, column=insert_at + i)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Populate data
            for result in results:
                row_num = result.row_index + 2  # +2 for header and 0-indexing
                extracted = result.extracted_data
                
                # UPDATED: Added min/max bill rate values
                values = [
                    extracted.bill_rate or "",
                    extracted.min_bill_rate if extracted.min_bill_rate is not None else "",  # NEW
                    extracted.max_bill_rate if extracted.max_bill_rate is not None else "",  # NEW
                    extracted.duration or "",
                    extracted.experience_required or "",
                    extracted.gbams_rgs_id or "",
                    extracted.ai_location or "",
                    ", ".join(extracted.skills) if extracted.skills else "",
                    extracted.role_description or "",
                    extracted.msp_owner or ""
                ]
                
                for i, value in enumerate(values):
                    cell = sheet.cell(row=row_num, column=insert_at + i)
                    cell.value = value
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output
            
        except Exception as e:
            logger.error(f"Failed to create result workbook: {e}")
            raise